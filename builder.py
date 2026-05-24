"""
builder.py
Builds the full Employee Effort Tracker workbook after consolidated_efforts is written.
Employee list and project codes are loaded from config/config.json.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ── Palette ───────────────────────────────────────────────────────────────────
HEADER_BG        = "2F5496"
HEADER_FONT      = "FFFFFF"
TOTALS_BG        = "D6DCE4"
SQUAD_COLORS     = {
    "Fighters": "DDEBF7",
    "Hero":     "E2EFDA",
    "Ninga":    "FFF2CC",
}
FREELANCER_BG    = "FCE4D6"
FREELANCER_FONT  = "C55A11"
INTERN_BG        = "E8D5F5"
INTERN_FONT      = "7030A0"
LOW_EFFORT_BG    = "FF9999"
LOW_EFFORT_THRESHOLD = 16   # 10% of 160 hrs


# ── Config loader ─────────────────────────────────────────────────────────────
_CONFIG_PATH = Path(__file__).parent / "config" / "config.json"

def load_config() -> dict:
    if not _CONFIG_PATH.exists():
        raise FileNotFoundError(f"Config file not found: {_CONFIG_PATH}")
    with open(_CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)

def get_lookup_data() -> list[tuple]:
    """Returns list of (name, code, contract, squad_lead, squad_name) from config."""
    cfg = load_config()
    return [
        (e["name"], e["code"], e["contract"], e["squad_lead"], e["squad_name"])
        for e in cfg.get("employees", [])
    ]

def get_project_codes() -> list[str]:
    cfg = load_config()
    return cfg.get("project_codes", [])


# ── Helpers ───────────────────────────────────────────────────────────────────
def _hstyle():
    return dict(
        font=Font(bold=True, color=HEADER_FONT),
        fill=PatternFill("solid", fgColor=HEADER_BG),
        alignment=Alignment(horizontal="center", vertical="center"),
    )

def _apply(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color, end_color=hex_color)


# ── Employee extraction ───────────────────────────────────────────────────────
def extract_employees(df: pd.DataFrame) -> list[str]:
    """Pull unique display names from 'Assigned To' col (format: Full Name <DOMAIN\\user>)."""
    if "Assigned To" not in df.columns:
        return []
    employees, seen = [], set()
    for val in df["Assigned To"].dropna():
        val = str(val).strip()
        name = val.split(" <")[0].strip() if " <" in val else val
        if name and name not in seen:
            seen.add(name)
            employees.append(name)
    return sorted(employees)


# ── Sheet: Lookup ─────────────────────────────────────────────────────────────
def _build_lookup(wb, lookup_data: list[tuple], new_employees: list[str]):
    if "Lookup" in wb.sheetnames:
        del wb["Lookup"]
    ws = wb.create_sheet("Lookup")

    headers = ["Employee Name", "Employee Code", "Employment Contract", "Squad Lead", "Squad Name"]
    widths  = [30, 16, 22, 25, 20]
    hs = _hstyle()
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws.cell(row=1, column=i, value=h), hs)
        ws.column_dimensions[get_column_letter(i)].width = w

    # Static rows from config
    for r, row in enumerate(lookup_data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # Append new employees found in data but not in config
    extra_start = len(lookup_data) + 2
    for r, name in enumerate(new_employees, extra_start):
        ws.cell(row=r, column=1, value=name)

    last_row = extra_start + len(new_employees) - 1 if new_employees else len(lookup_data) + 1

    # Dropdown: Employment Contract
    dv = DataValidation(type="list", formula1='"Full-time,Freelancer,Intern"',
                        allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv)
    dv.sqref = f"C2:C{max(last_row, 200)}"

    return ws


# ── Sheet: Project Code ───────────────────────────────────────────────────────
def _build_project_code(wb, project_codes: list[str]):
    if "Project Code" in wb.sheetnames:
        del wb["Project Code"]
    ws = wb.create_sheet("Project Code")

    _apply(ws.cell(row=1, column=1, value="Project Code"), _hstyle())
    ws.column_dimensions["A"].width = 20

    codes = project_codes if project_codes else ["PRJ-001"]
    for i, code in enumerate(codes, 2):
        ws.cell(row=i, column=1, value=code)

    tab = Table(displayName="ProjectCode", ref=f"A1:A{len(codes) + 1}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tab)
    return ws


# ── Sheet: Employees Effort ───────────────────────────────────────────────────
def _build_employees_effort(wb, employees: list[str], data_last_row: int, project_codes: list[str]):
    if "Employees Effort" in wb.sheetnames:
        del wb["Employees Effort"]
    ws = wb.create_sheet("Employees Effort")

    headers = ["Employee Name", "Employee Code", "Employment Contract",
               "Basic Effort", "Incentive Effort", "Total Effort",
               "Squad Lead", "Squad Name", "Project Code"]
    widths  = [30, 15, 22, 14, 16, 12, 25, 20, 15]
    hs = _hstyle()
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        _apply(ws.cell(row=1, column=i, value=h), hs)
        ws.column_dimensions[get_column_letter(i)].width = w

    n            = len(employees)
    ce_end       = data_last_row + 1
    last_emp_row = n + 1
    totals_row   = n + 2

    proj_dv_end  = len(project_codes) + 1 if project_codes else 4

    for r, emp in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp)
        ws.cell(row=r, column=2, value=f'=IFERROR(VLOOKUP(A{r},Lookup!$A:$B,2,0),"")')
        ws.cell(row=r, column=3, value=f'=IFERROR(VLOOKUP(A{r},Lookup!$A:$C,3,0),"")')
        ws.cell(row=r, column=4, value=(
            f"=SUMPRODUCT("
            f"(ISNUMBER(SEARCH(A{r},consolidated_efforts!$D$2:$D${ce_end})))"
            f"*(consolidated_efforts!$G$2:$G${ce_end}=0)"
            f"*consolidated_efforts!$H$2:$H${ce_end})"
        ))
        ws.cell(row=r, column=5, value=(
            f"=SUMPRODUCT("
            f"(ISNUMBER(SEARCH(A{r},consolidated_efforts!$D$2:$D${ce_end})))"
            f"*(consolidated_efforts!$G$2:$G${ce_end}=1)"
            f"*consolidated_efforts!$H$2:$H${ce_end})"
        ))
        ws.cell(row=r, column=6, value=f"=D{r}+E{r}")
        ws.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(A{r},Lookup!$A:$D,4,0),"")')
        ws.cell(row=r, column=8, value=f'=IFERROR(VLOOKUP(A{r},Lookup!$A:$E,5,0),"")')
        ws.cell(row=r, column=9, value="")

    # Project Code dropdown
    dv_proj = DataValidation(
        type="list",
        formula1=f"'Project Code'!$A$2:$A${proj_dv_end}",
        allow_blank=True,
    )
    ws.add_data_validation(dv_proj)
    dv_proj.sqref = f"I2:I{last_emp_row}"

    # ── Totals row ────────────────────────────────────────────────────────────
    grey_fill  = _fill(TOTALS_BG)
    bold_font  = Font(bold=True)
    top_border = Border(top=Side(style="medium", color="000000"))

    for col in range(1, 10):
        cell = ws.cell(row=totals_row, column=col)
        cell.fill   = grey_fill
        cell.font   = bold_font
        cell.border = top_border

    ws.cell(row=totals_row, column=1, value="Total")
    ws.cell(row=totals_row, column=4, value=f"=SUM(D2:D{last_emp_row})")
    ws.cell(row=totals_row, column=5, value=f"=SUM(E2:E{last_emp_row})")
    ws.cell(row=totals_row, column=6, value=f"=SUM(F2:F{last_emp_row})")

    # ── Conditional formatting ────────────────────────────────────────────────
    data_range = f"A2:I{last_emp_row}"
    c_range    = f"C2:C{last_emp_row}"

    # Priority: red (highest) → intern → freelancer → squad (lowest)
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f"$D2<{LOW_EFFORT_THRESHOLD}"],
        fill=_fill(LOW_EFFORT_BG), stopIfTrue=True,
    ))
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=['$C2="Intern"'], fill=_fill(INTERN_BG), stopIfTrue=False,
    ))
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=['$C2="Freelancer"'], fill=_fill(FREELANCER_BG), stopIfTrue=False,
    ))
    for squad, color in SQUAD_COLORS.items():
        ws.conditional_formatting.add(data_range, FormulaRule(
            formula=[f'$H2="{squad}"'], fill=_fill(color), stopIfTrue=False,
        ))
    ws.conditional_formatting.add(c_range, FormulaRule(
        formula=['$C2="Freelancer"'], font=Font(bold=True, color=FREELANCER_FONT),
    ))
    ws.conditional_formatting.add(c_range, FormulaRule(
        formula=['$C2="Intern"'], font=Font(bold=True, color=INTERN_FONT),
    ))

    ws.auto_filter.ref = f"A1:I{last_emp_row}"
    return ws


# ── Sheet ordering ────────────────────────────────────────────────────────────
def _reorder_sheets(wb, order: list[str]):
    existing = [s for s in order if s in wb.sheetnames]
    for name in reversed(existing):
        wb.move_sheet(name, offset=-wb.sheetnames.index(name))


# ── Public entry point ────────────────────────────────────────────────────────
def build_workbook(output_path: Path, combined_df: pd.DataFrame) -> dict:
    """
    Rebuilds all 4 sheets in the workbook.
    Returns a summary dict for the run report.
    """
    wb = load_workbook(output_path)

    lookup_data   = get_lookup_data()
    project_codes = get_project_codes()
    static_names  = {row[0] for row in lookup_data}

    names_in_data  = set(extract_employees(combined_df))
    new_employees  = sorted(e for e in names_in_data if e not in static_names)
    missing_from_data = [row[0] for row in lookup_data if row[0] not in names_in_data]

    all_employees  = [row[0] for row in lookup_data] + new_employees
    data_last_row  = len(combined_df)

    _build_lookup(wb, lookup_data, new_employees)
    _build_project_code(wb, project_codes)
    _build_employees_effort(wb, all_employees, data_last_row, project_codes)
    _reorder_sheets(wb, ["consolidated_efforts", "Employees Effort", "Lookup", "Project Code"])

    wb.save(output_path)

    return {
        "total_employees":     len(all_employees),
        "static_employees":    len(lookup_data),
        "new_employees":       new_employees,
        "missing_from_data":   missing_from_data,
        "project_codes":       project_codes,
        "data_rows":           data_last_row,
    }
