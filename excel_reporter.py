"""
excel_reporter.py
Exports the daily validation report as a formatted Excel file.
Rows are coloured per squad; violations appear as indented sub-rows.
"""

from __future__ import annotations
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from validator import EmployeeReport, Violation

# ── Palette ───────────────────────────────────────────────────────────────────
HEADER_BG        = "2F5496"
HEADER_FONT      = "FFFFFF"
SUMMARY_BG       = "D6DCE4"
SECTION_FONT     = "FFFFFF"

SQUAD_COLORS = {
    "Fighters": {"row": "DDEBF7", "header": "2E75B6"},
    "Hero":     {"row": "E2EFDA", "header": "375623"},
    "Ninga":    {"row": "FFF2CC", "header": "7F6000"},
}
DEFAULT_SQUAD = {"row": "F2F2F2", "header": "595959"}

CONTRACT_COLORS = {
    "Freelancer": "FCE4D6",
    "Intern":     "E8D5F5",
}

SEVERITY_COLORS = {
    "ERROR":   "FFD7D7",
    "WARNING": "FFF3CD",
    "INFO":    "D9EAF7",
}
SEVERITY_FONT = {
    "ERROR":   "C00000",
    "WARNING": "7F4F00",
    "INFO":    "0070C0",
}
CLEAN_COLOR  = "E2EFDA"
CLEAN_FONT   = "375623"
TOTALS_BG    = "D6DCE4"

COLS = {
    "name":          1,
    "contract":      2,
    "squad":         3,
    "squad_lead":    4,
    "basic_hrs":     5,
    "incentive_hrs": 6,
    "total_hrs":     7,
    "status":        8,
    "viol_code":     9,
    "viol_severity": 10,
    "viol_message":  11,
    "viol_detail":   12,
}
NUM_COLS = 12


# ── Helpers ───────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color, end_color=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(bold=bold, color=color, size=size)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _thin_border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _medium_bottom() -> Border:
    return Border(bottom=Side(style="medium", color="000000"))

def _set_row(ws: Worksheet, row: int, values: dict,
             fill: str | None = None, bold=False, font_color="000000",
             height: float | None = None):
    for col_key, val in values.items():
        col_idx = COLS[col_key]
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.alignment = _center() if col_key in ("basic_hrs","incentive_hrs","total_hrs",
                                                    "viol_code","viol_severity","status") else _left()
        cell.font  = _font(bold=bold, color=font_color)
        cell.border = _thin_border()
        if fill:
            cell.fill = _fill(fill)
    # Fill blank cells in the row so borders are consistent
    for col_idx in range(1, NUM_COLS + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is None:
            cell.border = _thin_border()
            if fill:
                cell.fill = _fill(fill)
    if height:
        ws.row_dimensions[row].height = height


# ── Sheet builders ────────────────────────────────────────────────────────────

def _write_title(ws: Worksheet, today: date) -> int:
    """Merged title + subtitle. Returns next free row."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    cell = ws.cell(row=1, column=1,
                   value=f"Daily Timesheet Validation Report — {today.strftime('%B %Y')}")
    cell.font      = Font(bold=True, color=HEADER_FONT, size=14)
    cell.fill      = _fill(HEADER_BG)
    cell.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    cell2 = ws.cell(row=2, column=1,
                    value=f"Generated: {today.strftime('%A, %d %B %Y')}")
    cell2.font      = Font(color="FFFFFF", size=10)
    cell2.fill      = _fill("1F3864")
    cell2.alignment = _center()
    ws.row_dimensions[2].height = 18
    return 3


def _write_summary(ws: Worksheet, reports: list[EmployeeReport], start_row: int) -> int:
    violated    = [r for r in reports if not r.is_clean]
    clean       = [r for r in reports if r.is_clean]
    error_count = sum(r.error_count   for r in reports)
    warn_count  = sum(r.warning_count for r in reports)

    # blank row
    ws.row_dimensions[start_row].height = 8
    r = start_row + 1

    # Summary header
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=1, value="Summary")
    cell.font      = Font(bold=True, color=HEADER_FONT, size=11)
    cell.fill      = _fill(HEADER_BG)
    cell.alignment = _left()
    ws.row_dimensions[r].height = 20
    r += 1

    summary_rows = [
        ("Total Employees",    len(reports)),
        ("Clean",              len(clean)),
        ("With Violations",    len(violated)),
        ("Hard Errors",        error_count),
        ("Warnings",           warn_count),
    ]
    for label, val in summary_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        lc = ws.cell(row=r, column=1, value=label)
        lc.font      = _font(bold=True)
        lc.fill      = _fill(SUMMARY_BG)
        lc.alignment = _left()
        lc.border    = _thin_border()
        vc = ws.cell(row=r, column=4, value=val)
        vc.font      = _font(bold=True)
        vc.fill      = _fill("FFFFFF")
        vc.alignment = _center()
        vc.border    = _thin_border()
        ws.row_dimensions[r].height = 16
        r += 1

    return r + 1   # blank row after summary


def _write_col_headers(ws: Worksheet, row: int) -> int:
    headers = [
        "Employee Name", "Contract", "Squad", "Squad Lead",
        "Basic Hrs", "Incentive Hrs", "Total Hrs", "Status",
        "Code", "Severity", "Violation Message", "Detail",
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font      = Font(bold=True, color=HEADER_FONT, size=10)
        cell.fill      = _fill(HEADER_BG)
        cell.alignment = _center()
        cell.border    = _thin_border()
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_squad_header(ws: Worksheet, row: int, squad_name: str,
                         squad_lead: str, squad_cfg: dict) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1,
                   value=f"  {squad_name}  —  Lead: {squad_lead}")
    cell.font      = Font(bold=True, color=SECTION_FONT, size=11)
    cell.fill      = _fill(squad_cfg["header"])
    cell.alignment = _left()
    for col_idx in range(1, NUM_COLS + 1):
        ws.cell(row=row, column=col_idx).border = _thin_border()
    ws.row_dimensions[row].height = 20
    return row + 1


def _status_label(r: EmployeeReport) -> str:
    if r.is_clean:
        return "CLEAN"
    parts = []
    if r.error_count:
        parts.append(f"{r.error_count} Error(s)")
    if r.warning_count:
        parts.append(f"{r.warning_count} Warning(s)")
    if r.violations and any(v.severity == "INFO" for v in r.violations):
        parts.append("Info")
    return " | ".join(parts)


def _write_employee_row(ws: Worksheet, row: int, report: EmployeeReport,
                         row_fill: str) -> int:
    """Write one employee summary row. Returns next row."""
    # Contract overrides squad color for background
    bg = CONTRACT_COLORS.get(report.contract, row_fill)

    status = _status_label(report)
    status_fill  = CLEAN_COLOR  if report.is_clean else row_fill
    status_fcolor = CLEAN_FONT  if report.is_clean else "C00000" if report.error_count else "7F4F00"

    _set_row(ws, row, {
        "name":          report.name,
        "contract":      report.contract,
        "squad":         report.squad_name,
        "squad_lead":    report.squad_lead,
        "basic_hrs":     round(report.total_basic_hrs, 1),
        "incentive_hrs": round(report.total_incentive_hrs, 1),
        "total_hrs":     round(report.total_hrs, 1),
        "status":        status,
    }, fill=bg, bold=True, height=18)

    # Override status cell colour
    sc = ws.cell(row=row, column=COLS["status"])
    sc.fill  = _fill(status_fill)
    sc.font  = _font(bold=True, color=status_fcolor)

    return row + 1


def _write_violation_row(ws: Worksheet, row: int, v: Violation,
                          squad_row_fill: str) -> int:
    """Write one violation sub-row under the employee."""
    vbg = SEVERITY_COLORS.get(v.severity, "FFFFFF")
    vfc = SEVERITY_FONT.get(v.severity, "000000")

    for col_idx in range(1, NUM_COLS + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill   = _fill(vbg)
        cell.border = _thin_border()
        cell.font   = _font(color=vfc, size=9)
        cell.alignment = _left()

    ws.cell(row=row, column=COLS["viol_code"],     value=v.code).alignment     = _center()
    ws.cell(row=row, column=COLS["viol_severity"], value=v.severity).alignment = _center()
    ws.cell(row=row, column=COLS["viol_message"],  value=v.message)
    if v.detail:
        ws.cell(row=row, column=COLS["viol_detail"], value=v.detail)

    ws.row_dimensions[row].height = 15
    return row + 1


def _write_totals(ws: Worksheet, row: int, reports: list[EmployeeReport]) -> int:
    total_basic     = sum(r.total_basic_hrs     for r in reports)
    total_incentive = sum(r.total_incentive_hrs for r in reports)
    total_all       = sum(r.total_hrs           for r in reports)

    _set_row(ws, row, {
        "name":          "TOTAL",
        "basic_hrs":     round(total_basic, 1),
        "incentive_hrs": round(total_incentive, 1),
        "total_hrs":     round(total_all, 1),
    }, fill=TOTALS_BG, bold=True, height=20)

    # Medium top border
    medium_top = Border(top=Side(style="medium", color="000000"))
    for col_idx in range(1, NUM_COLS + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.border = Border(
            top=Side(style="medium", color="000000"),
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        )
    return row + 1


def _set_col_widths(ws: Worksheet):
    widths = [28, 14, 12, 22, 11, 14, 11, 22, 8, 12, 45, 40]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ── Public entry point ────────────────────────────────────────────────────────

def save_report(reports: list[EmployeeReport], today: date, output_dir: str) -> Path:
    """Build and save the Excel report. Returns the output file path."""
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / f"report_{today.strftime('%Y-%m-%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"   # freeze title + summary area

    current_row = _write_title(ws, today)
    current_row = _write_summary(ws, reports, current_row)
    current_row = _write_col_headers(ws, current_row)

    # Group by squad
    squads: dict[str, list[EmployeeReport]] = {}
    for r in reports:
        squads.setdefault(r.squad_name, []).append(r)

    for squad_name, squad_reports in sorted(squads.items()):
        squad_cfg  = SQUAD_COLORS.get(squad_name, DEFAULT_SQUAD)
        row_fill   = squad_cfg["row"]
        squad_lead = squad_reports[0].squad_lead

        current_row = _write_squad_header(ws, current_row, squad_name, squad_lead, squad_cfg)

        for report in sorted(squad_reports, key=lambda r: r.name):
            current_row = _write_employee_row(ws, current_row, report, row_fill)
            for v in sorted(report.violations, key=lambda x: x.severity):
                current_row = _write_violation_row(ws, current_row, v, row_fill)

    current_row = _write_totals(ws, current_row, reports)

    _set_col_widths(ws)
    ws.auto_filter.ref = f"A{current_row - len(reports) - len(squads) - 2}:L{current_row - 1}"

    wb.save(output_path)
    return output_path
