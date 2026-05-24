"""
consolidator.py
Merges all downloaded CSV files into a dated copy of template.xlsx
saved under the output/ folder, then builds the full tracker workbook.
"""

import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from builder import build_workbook


def merge_csvs_into_template(
    csv_paths: list[Path],
    template_file: str,
    output_dir: str = "output",
) -> tuple[Path, dict]:
    """
    Reads all CSV files, writes them to the consolidated_efforts sheet of a
    dated copy of template.xlsx inside output_dir, then builds the full workbook.
    Returns (output_path, summary_dict).
    """
    if not csv_paths:
        raise ValueError("No CSV files to consolidate.")

    template_path = Path(template_file)
    if not template_path.exists():
        raise FileNotFoundError(
            f"Template file not found: {template_path}\n"
            "Place your Excel template at the configured TEMPLATE_FILE path."
        )

    # ── Load & combine CSVs ────────────────────────────────────────────────────
    frames = []
    for path in csv_paths:
        df = pd.read_csv(path, encoding="utf-8-sig")
        frames.append(df)

    combined = pd.concat(frames, ignore_index=True)
    combined = combined.where(pd.notnull(combined), None)

    # ── Dated output file in output/ ──────────────────────────────────────────
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp   = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_path = out_dir / f"consolidated_output_{timestamp}.xlsx"

    shutil.copy2(template_path, output_path)

    # ── Write consolidated_efforts sheet ──────────────────────────────────────
    wb = load_workbook(output_path)
    sheet_name = "consolidated_efforts"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    columns = list(combined.columns)
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_offset, (_, data_row) in enumerate(combined.iterrows()):
        target_row = 2 + row_offset
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=target_row, column=col_idx, value=data_row[col_name])
    wb.save(output_path)

    # ── Build full tracker (Employees Effort, Lookup, Project Code) ───────────
    summary = build_workbook(output_path, combined)
    summary["csv_files"]    = [p.name for p in csv_paths]
    summary["output_file"]  = str(output_path.resolve())
    summary["timestamp"]    = timestamp

    return output_path, summary
